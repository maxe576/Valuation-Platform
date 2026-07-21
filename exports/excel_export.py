"""Excel export of a valuation (§28).

Produces a multi-sheet workbook: Summary, Assumptions, Method Results, DCF
Projection, and a Disclaimer — stamped with company, date, analyst, model
version, and sources. Returns bytes so the UI can offer a download.
"""
from __future__ import annotations

import datetime as _dt
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from models.common import Scenario

_DISCLAIMER = (
    "This document is produced by a student-managed investment fund research "
    "platform for educational and research purposes only. It is NOT investment "
    "advice and NOT a recommendation to buy or sell any security. Valuations are "
    "estimates with material uncertainty. Verify all figures against primary "
    "filings before use."
)


def build_excel(
    company,
    fv,
    aset,
    analyst: str = "analyst",
    sources: Optional[list[str]] = None,
) -> bytes:
    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    _header(ws, "Valuation Summary")
    rows = [
        ("Company", f"{company.name} ({company.ticker})"),
        ("Valuation date", _dt.date.today().isoformat()),
        ("Analyst", analyst),
        ("Model version", aset.model_version),
        ("Sources", ", ".join(sources or [])),
        ("", ""),
        ("Current price", fv.current_price),
        ("Bear fair value", round(fv.bear, 2)),
        ("Base fair value", round(fv.base, 2)),
        ("Bull fair value", round(fv.bull, 2)),
        ("Blended fair value", round(fv.blended_value, 2)),
        ("Upside / (downside)", round(fv.upside, 4) if fv.upside is not None else None),
        ("Margin of safety",
         round(fv.margin_of_safety, 4) if fv.margin_of_safety is not None else None),
        ("Confidence score", fv.confidence_score),
        ("Method dispersion", round(fv.blend.dispersion, 4)),
    ]
    for label, value in rows:
        ws.append([label, value])

    # --- Assumptions ---
    wa = wb.create_sheet("Assumptions")
    _header(wa, "Forecast Assumptions")
    wa.append(["WACC", aset.wacc])
    wa.append(["Terminal growth", aset.terminal_growth])
    wa.append(["Exit EV/EBITDA", aset.exit_multiple])
    wa.append(["Cash", aset.cash])
    wa.append(["Total debt", aset.total_debt])
    wa.append(["Shares outstanding", aset.shares_outstanding])
    wa.append([])
    for scen in (Scenario.BEAR, Scenario.BASE, Scenario.BULL):
        sa = aset.scenarios.get(scen)
        if not sa:
            continue
        wa.append([scen.value.title()])
        wa.append(["  Year"] + [f"Y{i+1}" for i in range(sa.years())])
        wa.append(["  Revenue growth"] + list(sa.revenue_growth))
        wa.append(["  EBIT margin"] + list(sa.ebit_margin))
        wa.append(["  Tax rate"] + list(sa.tax_rate))
        wa.append([])

    # --- Method Results ---
    wm = wb.create_sheet("Method Results")
    _header(wm, "Method Results")
    wm.append(["Method", "Per share", "Template wt", "Normalized wt", "Contribution"])
    for c in fv.blend.contributions:
        wm.append([c.method.value, round(c.per_share_value, 2),
                   round(c.template_weight, 3), round(c.normalized_weight, 3),
                   round(c.weighted_value, 2)])

    # --- DCF Projection (base case) ---
    wd = wb.create_sheet("DCF Projection")
    _header(wd, "DCF — Base Case")
    wd.append(["Year", "Revenue", "EBIT", "Unlevered FCF", "PV of FCF"])
    for y in fv.dcf[Scenario.BASE].years:
        wd.append([y.year_index, round(y.revenue, 0), round(y.ebit, 0),
                   round(y.unlevered_fcf, 0), round(y.pv_unlevered_fcf, 0)])

    # --- Disclaimer ---
    wz = wb.create_sheet("Disclaimer")
    wz["A1"] = _DISCLAIMER
    wz["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    wz.column_dimensions["A"].width = 100

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header(ws, title: str) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
