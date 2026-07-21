"""Export tests: Excel bytes, CSV content, markdown memo (§29)."""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from models.company import Company
from config.lifecycle_weights import Lifecycle
from exports.csv_export import method_results_csv, valuation_summary_csv
from exports.excel_export import build_excel
from exports.memo_export import build_markdown_memo
from services.valuation_service import build_demo_valuation_inputs, run_full_valuation


def _fixtures():
    inputs = build_demo_valuation_inputs()
    fv = run_full_valuation(inputs)
    company = Company(ticker="NFLX", name="Netflix, Inc.",
                      lifecycle=Lifecycle.HIGH_GROWTH_PROFITABLE)
    return company, fv, inputs.assumption_set


def test_excel_export_is_valid_workbook():
    company, fv, aset = _fixtures()
    data = build_excel(company, fv, aset, analyst="max", sources=["SEC EDGAR"])
    assert isinstance(data, bytes) and len(data) > 0
    wb = load_workbook(BytesIO(data))
    assert set(["Summary", "Assumptions", "Method Results",
                "DCF Projection", "Disclaimer"]).issubset(set(wb.sheetnames))
    # Summary carries the blended value somewhere.
    summary_vals = [c.value for row in wb["Summary"].iter_rows() for c in row]
    assert "Blended fair value" in summary_vals


def test_csv_summary_contains_key_fields():
    company, fv, _ = _fixtures()
    csv_text = valuation_summary_csv(company, fv, sources=["SEC EDGAR"])
    assert "blended" in csv_text
    assert "DISCLAIMER" in csv_text
    assert method_results_csv(fv).startswith("method,per_share")


def test_markdown_memo_includes_sections():
    company, fv, _ = _fixtures()
    md = build_markdown_memo(company, fv, ai_output={
        "executive_summary": "Test summary.",
        "risks": ["Risk A", "Risk B"],
    })
    assert "# Netflix, Inc. (NFLX)" in md
    assert "Valuation summary" in md
    assert "Risk A" in md
    assert "not investment advice" in md
